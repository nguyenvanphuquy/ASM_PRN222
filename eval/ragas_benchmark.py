#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RAGAS benchmark cho ChatBot PRN222 — đo 4 chỉ số RAG chuẩn RAGAS:

  • faithfulness       : câu trả lời có bám vào ngữ cảnh truy xuất không (không bịa)
  • answer_relevancy   : câu trả lời có sát câu hỏi không
  • context_precision  : các đoạn truy xuất có liên quan & xếp hạng tốt không
  • context_recall     : ngữ cảnh truy xuất có đủ để suy ra ground-truth không

Cài đặt theo đúng ĐỊNH NGHĨA của RAGAS (https://docs.ragas.io) nhưng KHÔNG dùng thư viện
`ragas` (phụ thuộc langchain chồng chéo, hay vỡ trên Python mới). Thay vào đó:
  - Groq làm "judge" LLM (mặc định llama-3.1-8b-instant cho nhẹ rate-limit; sinh câu trả lời
    dùng đúng model của chatbot: llama-3.3-70b)
  - sentence-transformers (multilingual-e5-base) chạy local → cosine cho answer_relevancy
Không cần OpenAI key.

Chạy:  python eval/ragas_benchmark.py [--limit 10] [--judge llama-3.1-8b-instant]
Cài:   pip install -r eval/ragas_requirements.txt
"""
import argparse
import json
import os
import re
import sys
import time
import unicodedata

import numpy as np
import pyodbc
import requests
from openpyxl import Workbook, load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
APPSETTINGS = os.path.join(ROOT, "PresentationLayer", "appsettings.json")
TESTSET = os.path.join(ROOT, "TestSet_50cau_GroundTruth.xlsx")

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

GROQ_KEY = GEN_MODEL = JUDGE_MODEL = GROQ_URL = None
_last_call = [0.0]  # pacing giữa các lệnh Groq
PACE = [2.5]        # giây tối thiểu giữa 2 lệnh Groq (đặt cao cho model 70b để né TPM)


# ───────────────────────── Config ─────────────────────────
def load_config():
    with open(APPSETTINGS, encoding="utf-8-sig") as f:
        cfg = json.load(f)
    g = cfg["Groq"]
    # Cho phép override Groq key qua biến môi trường GROQ_API_KEY (dùng key khác nếu key chính hết quota ngày).
    key = os.environ.get("GROQ_API_KEY", "").strip() or g["ApiKey"]
    return (key, g.get("Model", "llama-3.3-70b-versatile"),
            g.get("BaseUrl", "https://api.groq.com/openai/v1"),
            cfg["ConnectionStrings"]["DefaultConnection"])


def parse_conn(s):
    d = {}
    for part in s.split(";"):
        if "=" in part:
            k, v = part.split("=", 1)
            d[k.strip().lower()] = v.strip()
    return d.get("server", "localhost"), d.get("database", "ChatBotPRN222"), d.get("user id", "admin"), d.get("password", "123")


def db_connect(s):
    server, database, uid, pwd = parse_conn(s)
    for drv in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server", "SQL Server"):
        try:
            return pyodbc.connect(f"DRIVER={{{drv}}};SERVER={server};DATABASE={database};UID={uid};PWD={pwd};TrustServerCertificate=yes", timeout=5)
        except Exception:
            continue
    raise RuntimeError("Không kết nối được SQL Server (kiểm tra ODBC Driver).")


# ───────────────────────── Groq ─────────────────────────
def _retry_after(r):
    """Đọc thời gian chờ Groq yêu cầu (header Retry-After hoặc 'try again in 1m2.3s')."""
    h = r.headers.get("retry-after")
    if h:
        try:
            return float(h) + 1
        except Exception:
            pass
    m = re.search(r"try again in\s*(?:(\d+)m)?\s*([\d.]+)s", r.text)
    if m:
        return (int(m.group(1) or 0) * 60 + float(m.group(2))) + 1
    return None


def groq_chat(messages, model, temperature=0.0, max_tokens=500, pace=None):
    # pace tối thiểu giữa 2 lệnh để không vượt token/phút (TPM) của Groq free-tier
    pace = PACE[0] if pace is None else pace
    dt = time.time() - _last_call[0]
    if dt < pace:
        time.sleep(pace - dt)
    body = {"model": model, "messages": messages, "temperature": temperature, "max_tokens": max_tokens}
    for attempt in range(6):
        r = requests.post(f"{GROQ_URL}/chat/completions",
                          headers={"Authorization": f"Bearer {GROQ_KEY}"}, json=body, timeout=90)
        _last_call[0] = time.time()
        if r.status_code == 429 and attempt < 5:
            wait = _retry_after(r) or (10 * (attempt + 1))
            print(f"    (429 → chờ {wait:.0f}s)", flush=True)
            time.sleep(min(wait, 75))
            _last_call[0] = time.time()
            continue
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"].strip()
    r.raise_for_status()


def judge_json(prompt):
    txt = groq_chat([{"role": "user", "content": prompt}], JUDGE_MODEL, max_tokens=500)
    m = re.search(r"\{.*\}|\[.*\]", txt, re.S)
    try:
        return json.loads(m.group(0) if m else txt)
    except Exception:
        return None


def _yes(x):
    return 1 if str(x).strip().lower() in ("1", "true", "yes", "có", "co") else 0


# ───────────────────────── Retrieval (keyword, giống app) ─────────────────────────
STOP = set("tom tat cho toi tai lieu cua mon hay nhe voi nay minh ban oi duoc the nao lai di va cac mot "
           "nhung trong ve la co khong hoac thi se day dum list ke ra noi gi nhu de khi".split())


def no_accents(s):
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D").lower()


def retrieve(cur, subject_id, query, limit=5):
    cur.execute("SELECT Content, DocumentName, Page FROM DocumentChunks WHERE SubjectId=?", subject_id)
    chunks = [(r[0], r[1], r[2]) for r in cur.fetchall()]
    if not chunks:
        return []
    terms = {t for t in re.split(r"[ ,.?!:;()\[\]\"'/\\*#]+", no_accents(query)) if len(t) >= 3 and t not in STOP}
    scored = []
    for content, name, page in chunks:
        cn = no_accents(content)
        hit = sum(1 for t in terms if t in cn)
        if hit:
            scored.append((hit / max(1, len(terms)), content))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [c for _, c in scored[:limit]]


RAG_SYS = ("Bạn là trợ lý học tập AI, trả lời bằng tiếng Việt. Quy tắc: CHỈ trả lời dựa trên NGỮ CẢNH "
           "tài liệu bên dưới; nếu không đủ thông tin hãy trả lời đúng câu: "
           "\"Tôi không tìm thấy thông tin này trong tài liệu môn học.\" Ngắn gọn, không bịa.")


def rag_answer(question, contexts):
    if not contexts:
        return "Tôi không tìm thấy thông tin này trong tài liệu môn học."
    ctx = "\n\n".join(f"[{i+1}] {c}" for i, c in enumerate(contexts))
    return groq_chat([{"role": "system", "content": f"{RAG_SYS}\n\n=== NGỮ CẢNH ===\n{ctx}\n=== HẾT ==="},
                      {"role": "user", "content": question}], GEN_MODEL, max_tokens=500)


# ───────────────────────── 4 chỉ số RAGAS ─────────────────────────
def m_faithfulness(answer, contexts):
    if "không tìm thấy thông tin" in answer.lower() or not contexts:
        return 0.0
    claims = judge_json("Tách câu trả lời sau thành danh sách các khẳng định NGẮN, độc lập. "
                        f"Chỉ trả JSON mảng chuỗi.\nCâu trả lời: {answer}")
    if not isinstance(claims, list) or not claims:
        return None
    ctx = "\n".join(contexts)
    v = judge_json("Nhiệm vụ: với NGỮ CẢNH cho trước, chấm mỗi khẳng định là 1 nếu Ý của nó ĐƯỢC NGỮ CẢNH "
                   "nói tới/hỗ trợ (dù diễn đạt khác), 0 nếu ngữ cảnh KHÔNG đề cập. "
                   "CHỈ trả về một mảng JSON gồm các số 0/1, độ dài đúng bằng số khẳng định, không giải thích.\n"
                   f"NGỮ CẢNH:\n{ctx}\n\nDANH SÁCH KHẲNG ĐỊNH (JSON):\n" + json.dumps(claims, ensure_ascii=False))
    if not isinstance(v, list) or not v:
        return None
    vv = [_yes(x) for x in v]
    return sum(vv) / len(vv)


def m_answer_relevancy(question, answer, embed):
    if "không tìm thấy thông tin" in answer.lower():
        return 0.0
    gen = judge_json("Từ CÂU TRẢ LỜI sau, tạo 3 câu hỏi mà câu trả lời này giải đáp. Chỉ trả JSON mảng 3 chuỗi.\n"
                     f"CÂU TRẢ LỜI: {answer}")
    if not isinstance(gen, list) or not gen:
        return None
    vecs = embed([question] + [str(x) for x in gen][:3])
    q0 = vecs[0]
    sims = [float(np.dot(q0, v) / (np.linalg.norm(q0) * np.linalg.norm(v) + 1e-9)) for v in vecs[1:]]
    return max(0.0, sum(sims) / len(sims)) if sims else None


def m_context_precision(contexts, ground_truth):
    if not contexts:
        return 0.0
    listed = "\n".join(f"[{i+1}] {c}" for i, c in enumerate(contexts))
    v = judge_json("Nhiệm vụ: với CÂU TRẢ LỜI ĐÚNG cho trước, chấm mỗi ĐOẠN là 1 nếu đoạn đó CHỨA thông tin "
                   "hữu ích để tạo ra câu trả lời đúng, 0 nếu không liên quan. "
                   "CHỈ trả về mảng JSON các số 0/1, độ dài đúng bằng số đoạn, không giải thích.\n"
                   f"CÂU TRẢ LỜI ĐÚNG: {ground_truth}\n\nCÁC ĐOẠN:\n{listed}")
    if not isinstance(v, list) or not v:
        return None
    rel = [_yes(x) for x in v][:len(contexts)]
    hit, precisions = 0, []
    for i, r in enumerate(rel):
        if r:
            hit += 1
            precisions.append(hit / (i + 1))
    return sum(precisions) / hit if hit else 0.0


def m_context_recall(contexts, ground_truth):
    stmts = judge_json("Tách CÂU TRẢ LỜI ĐÚNG sau thành danh sách các ý NGẮN, độc lập. Chỉ trả JSON mảng chuỗi.\n"
                       f"CÂU TRẢ LỜI ĐÚNG: {ground_truth}")
    if not isinstance(stmts, list) or not stmts:
        return None
    ctx = "\n".join(contexts) if contexts else "(rỗng)"
    v = judge_json("Nhiệm vụ: với NGỮ CẢNH cho trước, chấm mỗi Ý là 1 nếu Ý đó có thể suy ra/được NGỮ CẢNH "
                   "hỗ trợ (dù diễn đạt khác), 0 nếu ngữ cảnh không đề cập. "
                   "CHỈ trả về mảng JSON các số 0/1, độ dài đúng bằng số ý, không giải thích.\n"
                   f"NGỮ CẢNH:\n{ctx}\n\nCÁC Ý (JSON):\n" + json.dumps(stmts, ensure_ascii=False))
    if not isinstance(v, list) or not v:
        return None
    vv = [_yes(x) for x in v]
    return sum(vv) / len(vv)


# ───────────────────────── Test set ─────────────────────────
def load_testset():
    ws = load_workbook(TESTSET).active
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows) if r and any(str(c).strip().upper() == "STT" for c in r if c))
    items = []
    for r in rows[hi + 1:]:
        if not r or r[0] in (None, ""):
            continue
        items.append({"stt": r[0], "subject": str(r[1]).strip() if r[1] else "",
                      "topic": str(r[2]).strip() if len(r) > 2 and r[2] else "",
                      "question": str(r[3]).strip() if len(r) > 3 and r[3] else "",
                      "ground_truth": str(r[4]).strip() if len(r) > 4 and r[4] else ""})
    return items


# ───────────────────────── Output ─────────────────────────
def save_outputs(results):
    metrics = ["faithfulness", "answer_relevancy", "context_precision", "context_recall"]

    def avg(k):
        xs = [r[k] for r in results if r.get(k) is not None]
        return sum(xs) / len(xs) if xs else float("nan")

    wb = Workbook(); ws = wb.active; ws.title = "RAGAS"
    ws.append(["STT", "Môn", "Câu hỏi", "#ctx"] + metrics)
    for r in results:
        ws.append([r["stt"], r["subject"], r["question"], r["n_ctx"]] +
                  [round(r[m], 3) if r.get(m) is not None else "" for m in metrics])
    ws.append([]); ws.append(["", "", "TRUNG BÌNH", ""] + [round(avg(m), 3) for m in metrics])
    xlsx = os.path.join(HERE, "RAGAS_Results.xlsx"); wb.save(xlsx)

    md = ["# Kết quả RAGAS benchmark", "",
          f"- Sinh câu trả lời: `{GEN_MODEL}` · Judge: `{JUDGE_MODEL}` (Groq) · Embedding: `multilingual-e5-base` (local)",
          f"- Số câu đánh giá: **{len(results)}** (thuộc môn đã có tài liệu index)", "",
          "## Trung bình", "", "| Chỉ số | Giá trị |", "|---|---|"]
    for m in metrics:
        md.append(f"| {m} | **{avg(m):.3f}** |")
    md += ["", "## Chi tiết từng câu", "",
           "| STT | Môn | #ctx | faithfulness | answer_relevancy | context_precision | context_recall |",
           "|---|---|---|---|---|---|---|"]
    for r in results:
        md.append("| {} | {} | {} | {} |".format(r["stt"], r["subject"], r["n_ctx"],
                  " | ".join(f"{r[m]:.3f}" if r.get(m) is not None else "—" for m in metrics)))
    with open(os.path.join(HERE, "RAGAS_Results.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(md) + "\n")

    print("\n=== KẾT QUẢ RAGAS (trung bình) ===")
    for m in metrics:
        print(f"  {m:20s}: {avg(m):.3f}")
    print(f"\nĐã lưu: {xlsx}\n         {os.path.join(HERE, 'RAGAS_Results.md')}")


# ───────────────────────── Main ─────────────────────────
def main():
    global GROQ_KEY, GEN_MODEL, JUDGE_MODEL, GROQ_URL
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=10)
    ap.add_argument("--gen", default="", help="model Groq sinh câu trả lời (mặc định = model chatbot)")
    ap.add_argument("--judge", default="", help="model Groq để chấm (mặc định = model sinh)")
    ap.add_argument("--pace", type=float, default=2.5, help="giây giữa 2 lệnh Groq (đặt ~10 cho 70b)")
    args = ap.parse_args()
    PACE[0] = args.pace

    GROQ_KEY, GEN_MODEL, GROQ_URL, conn_str = load_config()
    if args.gen:
        GEN_MODEL = args.gen
    JUDGE_MODEL = args.judge or GEN_MODEL
    if not GROQ_KEY or not GROQ_KEY.startswith("gsk_"):
        sys.exit("Thiếu Groq API key hợp lệ trong appsettings.json.")

    print("Nạp embedding model (multilingual-e5-base) ...", flush=True)
    from sentence_transformers import SentenceTransformer
    st = SentenceTransformer("intfloat/multilingual-e5-base", device="cpu")
    def embed(texts):
        return st.encode(["query: " + t for t in texts], normalize_embeddings=True)

    cn = db_connect(conn_str); cur = cn.cursor()
    cur.execute("SELECT s.Id, s.Code FROM Subjects s WHERE EXISTS (SELECT 1 FROM DocumentChunks c WHERE c.SubjectId=s.Id)")
    subj = {code.upper(): sid for sid, code in cur.fetchall()}
    print("Môn có tài liệu:", list(subj.keys()), flush=True)

    items = [it for it in load_testset() if it["subject"].upper() in subj and it["question"] and it["ground_truth"]]
    print(f"Câu hỏi (thuộc môn có tài liệu): {len(items)} — chạy tối đa {args.limit}.", flush=True)
    items = items[:args.limit]
    if not items:
        sys.exit("Không có câu hỏi nào thuộc môn đã có tài liệu index.")

    metrics = ["faithfulness", "answer_relevancy", "context_precision", "context_recall"]
    results = []
    try:
        for i, it in enumerate(items, 1):
            sid = subj[it["subject"].upper()]
            contexts = retrieve(cur, sid, it["question"], limit=5)
            row = {**it, "n_ctx": len(contexts), "answer": ""}
            try:
                answer = rag_answer(it["question"], contexts)
                row["answer"] = answer
                row["faithfulness"] = m_faithfulness(answer, contexts)
                row["answer_relevancy"] = m_answer_relevancy(it["question"], answer, embed)
                row["context_precision"] = m_context_precision(contexts, it["ground_truth"])
                row["context_recall"] = m_context_recall(contexts, it["ground_truth"])
                print(f"[{i}/{len(items)}] {it['subject']} · {it['question'][:46]}... "
                      f"F={row['faithfulness']} R={row['answer_relevancy']} P={row['context_precision']} Rc={row['context_recall']}", flush=True)
            except Exception as ex:
                print(f"[{i}/{len(items)}] LỖI (giữ kết quả đã có): {ex}", flush=True)
                for m in metrics:
                    row.setdefault(m, None)
                results.append(row)
                break  # thường là hết quota → dừng, xuất phần đã chấm
            results.append(row)
            time.sleep(1.5)
    finally:
        if results:
            save_outputs(results)
        else:
            print("Không có kết quả nào để lưu.")


if __name__ == "__main__":
    main()
