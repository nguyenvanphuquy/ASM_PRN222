#!/usr/bin/env python3
"""
Chấm điểm chatbot bằng test set 50 câu hỏi + ground truth.

Quy trình:
  1) Đăng nhập web app (cookie auth) bằng tài khoản student.
  2) Tạo 1 phiên chat (không gắn môn → tìm trên toàn bộ tài liệu).
  3) Gửi từng câu hỏi tới /Chat/Ask, lấy câu trả lời của bot.
  4) Dùng Groq (LLM-as-judge) so câu trả lời với ground truth → CORRECT/INCORRECT.
     Nếu không có Groq key / mất mạng → fallback chấm bằng độ trùng từ khóa.
  5) In độ chính xác (accuracy) và xuất file Eval_Results.xlsx.

Yêu cầu: app đang chạy (dotnet run). Cài thư viện: pip install -r requirements.txt

Ví dụ:
  python evaluate_chatbot.py --base-url http://localhost:5217
  python evaluate_chatbot.py --base-url https://localhost:7189 --username student --password student123
"""
import argparse
import json
import os
import re
import sys
import time
import urllib.parse

import requests
import urllib3
from openpyxl import load_workbook, Workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(HERE, "..", "TestSet_50cau_GroundTruth.xlsx")
DEFAULT_APPSETTINGS = os.path.join(HERE, "..", "ChatBotPRN222", "appsettings.json")


# ----------------------------------------------------------------------------- read test set
def read_test_set(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # Tìm dòng header chứa "STT"
    header_idx = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "STT")
    items = []
    for r in rows[header_idx + 1:]:
        if not r or r[0] is None or str(r[0]).strip() == "":
            continue
        items.append({
            "stt": r[0],
            "subject": (r[1] or "").strip() if len(r) > 1 else "",
            "topic": (r[2] or "").strip() if len(r) > 2 else "",
            "question": (r[3] or "").strip() if len(r) > 3 else "",
            "ground_truth": (r[4] or "").strip() if len(r) > 4 else "",
            "source": (r[5] or "").strip() if len(r) > 5 else "",
        })
    return items


# ----------------------------------------------------------------------------- chatbot client
class ChatClient:
    def __init__(self, base_url):
        self.base = base_url.rstrip("/")
        self.s = requests.Session()
        self.s.verify = False
        self.s.headers.update({"User-Agent": "eval-script"})

    def login(self, username, password):
        # Lấy antiforgery token từ form login.
        r = self.s.get(f"{self.base}/Auth/Login", timeout=30)
        r.raise_for_status()
        m = re.search(r'name="__RequestVerificationToken"[^>]*value="([^"]+)"', r.text)
        if not m:
            raise RuntimeError("Không tìm thấy antiforgery token ở trang Login.")
        token = m.group(1)
        r = self.s.post(f"{self.base}/Auth/Login", data={
            "Username": username,
            "Password": password,
            "RememberMe": "false",
            "__RequestVerificationToken": token,
        }, timeout=30, allow_redirects=True)
        # Đăng nhập thành công → bị redirect khỏi trang Login.
        if "/Auth/Login" in r.url:
            raise RuntimeError("Đăng nhập thất bại — kiểm tra username/password.")

    def new_session(self, subject_id=None):
        r = self.s.post(f"{self.base}/Chat/NewSession",
                        json={"SubjectId": subject_id}, timeout=30)
        r.raise_for_status()
        return r.json()["id"]

    def ask(self, session_id, question):
        r = self.s.post(f"{self.base}/Chat/Ask",
                        json={"SessionId": session_id, "Question": question}, timeout=120)
        r.raise_for_status()
        data = r.json()
        sources = "; ".join(
            f"{s.get('documentName','?')} (p.{s.get('page','?')})" for s in data.get("sources", [])
        )
        return data.get("answer", ""), sources


# ----------------------------------------------------------------------------- judge
def load_groq(appsettings, override_key=None):
    cfg = {"ApiKey": override_key or "", "Model": "llama-3.3-70b-versatile",
           "BaseUrl": "https://api.groq.com/openai/v1"}
    if os.path.exists(appsettings):
        try:
            with open(appsettings, encoding="utf-8-sig") as f:
                g = json.load(f).get("Groq", {})
            cfg["Model"] = g.get("Model", cfg["Model"])
            cfg["BaseUrl"] = g.get("BaseUrl", cfg["BaseUrl"])
            if not override_key:
                cfg["ApiKey"] = g.get("ApiKey", "")
        except Exception:
            pass
    return cfg


def judge_llm(groq, question, ground_truth, answer):
    prompt = (
        "Bạn là giám khảo chấm điểm câu trả lời của một chatbot học tập.\n"
        "So sánh CÂU TRẢ LỜI với ĐÁP ÁN ĐÚNG (ground truth).\n"
        "Nếu câu trả lời truyền tải đúng ý chính của đáp án (dù diễn đạt khác) → CORRECT.\n"
        "Nếu sai, thiếu ý chính, hoặc bịa thông tin → INCORRECT.\n"
        "Trả về đúng 1 từ ở dòng đầu: CORRECT hoặc INCORRECT, dòng sau là 1 câu lý do ngắn.\n\n"
        f"CÂU HỎI: {question}\n\nĐÁP ÁN ĐÚNG: {ground_truth}\n\nCÂU TRẢ LỜI: {answer}\n"
    )
    r = requests.post(
        f"{groq['BaseUrl']}/chat/completions",
        headers={"Authorization": f"Bearer {groq['ApiKey']}", "Content-Type": "application/json"},
        json={"model": groq["Model"], "temperature": 0,
              "messages": [{"role": "user", "content": prompt}]},
        timeout=60,
    )
    r.raise_for_status()
    text = r.json()["choices"][0]["message"]["content"].strip()
    verdict = "Đúng" if text.upper().startswith("CORRECT") else "Sai"
    reason = text.split("\n", 1)[1].strip() if "\n" in text else ""
    return verdict, reason


def judge_overlap(ground_truth, answer):
    # Fallback đơn giản: tỉ lệ từ khóa (≥4 ký tự) của ground truth xuất hiện trong câu trả lời.
    def toks(s):
        return set(w for w in re.findall(r"\w+", s.lower()) if len(w) >= 4)
    gt, ans = toks(ground_truth), toks(answer)
    if not gt:
        return "Sai", "ground truth rỗng"
    ratio = len(gt & ans) / len(gt)
    return ("Đúng" if ratio >= 0.5 else "Sai"), f"trùng {ratio:.0%} từ khóa"


# ----------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--base-url", default="http://localhost:5217")
    ap.add_argument("--username", default="student")
    ap.add_argument("--password", default="student123")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--appsettings", default=DEFAULT_APPSETTINGS)
    ap.add_argument("--groq-key", default=None, help="Ghi đè Groq API key (mặc định đọc từ appsettings.json)")
    ap.add_argument("--no-judge", action="store_true", help="Không dùng LLM, chỉ chấm bằng trùng từ khóa")
    ap.add_argument("--out", default=os.path.join(HERE, "Eval_Results.xlsx"))
    ap.add_argument("--limit", type=int, default=0, help="Chỉ chạy N câu đầu (để thử nhanh)")
    args = ap.parse_args()

    items = read_test_set(args.xlsx)
    if args.limit:
        items = items[:args.limit]
    print(f"Đã đọc {len(items)} câu hỏi từ {os.path.basename(args.xlsx)}")

    groq = load_groq(args.appsettings, args.groq_key)
    use_llm = not args.no_judge and bool(groq["ApiKey"])
    print(f"Chấm điểm: {'Groq LLM-as-judge (' + groq['Model'] + ')' if use_llm else 'trùng từ khóa (fallback)'}")

    client = ChatClient(args.base_url)
    print(f"Đăng nhập {args.base_url} ...")
    client.login(args.username, args.password)
    session_id = client.new_session(subject_id=None)
    print(f"Đã tạo phiên chat: {session_id}\n")

    results = []
    correct = 0
    for it in items:
        try:
            answer, sources = client.ask(session_id, it["question"])
        except Exception as e:
            answer, sources = f"[LỖI gọi chatbot: {e}]", ""

        if use_llm:
            try:
                verdict, reason = judge_llm(groq, it["question"], it["ground_truth"], answer)
            except Exception as e:
                verdict, reason = judge_overlap(it["ground_truth"], answer)
                reason += f" (LLM judge lỗi: {e})"
        else:
            verdict, reason = judge_overlap(it["ground_truth"], answer)

        if verdict == "Đúng":
            correct += 1
        results.append({**it, "answer": answer, "sources": sources,
                        "verdict": verdict, "reason": reason})
        print(f"  #{it['stt']:>2} [{verdict}] {it['question'][:60]}")
        time.sleep(0.3)  # tránh spam API

    acc = correct / len(items) if items else 0
    print(f"\n=== KẾT QUẢ: {correct}/{len(items)} câu đúng — Accuracy = {acc:.1%} ===")

    # Xuất Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ket qua danh gia"
    headers = ["STT", "Mã môn", "Chủ đề", "Câu hỏi", "Đáp án đúng (Ground Truth)",
               "Câu trả lời của Bot", "Nguồn (chunks)", "Kết quả", "Lý do"]
    ws.append(headers)
    for r in results:
        ws.append([r["stt"], r["subject"], r["topic"], r["question"], r["ground_truth"],
                   r["answer"], r["sources"], r["verdict"], r["reason"]])
    ws.append([])
    ws.append(["", "", "", "", "", "", "Accuracy", f"{acc:.1%}", f"{correct}/{len(items)}"])
    wb.save(args.out)
    print(f"Đã ghi chi tiết vào: {args.out}")
    return 0 if acc > 0 else 1


if __name__ == "__main__":
    sys.exit(main())
